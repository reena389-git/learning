"""
DDL generator — reads a filled-in UC Metadata Enrichment workbook
and emits the Databricks SQL needed to apply the metadata.

Reads:
  - Table-level sheet:  one row per metadata field, column C = value
  - Column-level sheet: one row per column, columns A..R = the 18 metadata fields

Produces an ordered set of statements:
  1. Table comment (TBLPROPERTIES)
  2. Table tags (SET TAGS)
  3. Constraints — PK, FK, CHECK
  4. Column comments
  5. Column tags
"""

import sys
from openpyxl import load_workbook
from pathlib import Path

# ----------------------------------------------------------------------
# Column-level field order, MUST match the workbook column order
# ----------------------------------------------------------------------
COL_FIELDS = [
    "column_name",          # A
    "data_type",            # B  (informational only)
    "nullable",             # C
    "comment",              # D
    "business_name",        # E
    "role",                 # F
    "is_primary_key",       # G
    "is_foreign_key",       # H
    "fk_references",        # I
    "fk_cardinality",       # J
    "referential_integrity",# K
    "allowed_values",       # L
    "format_regex",         # M
    "source_of_truth",      # N
    "measure_grain",        # O
    "measure_unit",         # P
    "measure_methodology",  # Q
    "agg_rules",            # R
]


def normalise(val):
    """Treat blanks, None, 'n/a', '-' as empty. Also catches 'n/a — explanation' strings."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("n/a", "na", "-", "none", ""):
        return ""
    # Values that start with 'n/a' followed by an em-dash or hyphen are also "not applicable"
    sl = s.lower()
    if sl.startswith("n/a ") or sl.startswith("n/a—") or sl.startswith("n/a -") or sl.startswith("n/a-"):
        return ""
    return s


# ----------------------------------------------------------------------
# Reading the workbook
# ----------------------------------------------------------------------
def read_table_metadata(ws):
    """Read the table-level sheet — return a dict of field -> value."""
    meta = {}
    # Header row is at row 6 (after title block); data starts at row 7.
    # Find the header row dynamically by scanning for 'Metadata field' in column A.
    header_row = None
    for r in range(1, 30):
        v = ws.cell(row=r, column=1).value
        if v and "metadata field" in str(v).lower():
            header_row = r
            break
    if not header_row:
        raise RuntimeError("Could not find 'Metadata field' header in table-level sheet.")

    for r in range(header_row + 1, ws.max_row + 1):
        field = normalise(ws.cell(row=r, column=1).value)
        value = normalise(ws.cell(row=r, column=3).value)
        if field:
            meta[field.lower()] = value
    return meta


def read_column_metadata(ws):
    """Read the column-level sheet — return a list of dicts."""
    # Header is at the row where A1 says 'Column name'.
    header_row = None
    for r in range(1, 30):
        v = ws.cell(row=r, column=1).value
        if v and "column name" in str(v).lower():
            header_row = r
            break
    if not header_row:
        raise RuntimeError("Could not find 'Column name' header in column-level sheet.")

    # Data starts two rows below (subheader takes one row).
    data_start = header_row + 2

    columns = []
    for r in range(data_start, ws.max_row + 1):
        row = {}
        for i, field in enumerate(COL_FIELDS, start=1):
            row[field] = normalise(ws.cell(row=r, column=i).value)
        if row["column_name"]:
            columns.append(row)
    return columns


# ----------------------------------------------------------------------
# DDL generation
# ----------------------------------------------------------------------
def gen_ddl(fqn, table_meta, columns):
    """Generate the full DDL for one table."""
    lines = []
    indent = "  "

    lines.append(f"-- " + "=" * 70)
    lines.append(f"-- DDL for {fqn}")
    lines.append(f"-- Generated from UC Metadata Enrichment workbook")
    lines.append(f"-- " + "=" * 70)
    lines.append("")

    # ----------------------------------------------------------------
    # 1. Table comment
    # ----------------------------------------------------------------
    comment = table_meta.get("table comment (business description)", "")
    if comment:
        lines.append("-- 1. Table comment")
        lines.append(f"ALTER TABLE {fqn}")
        lines.append(f"  SET TBLPROPERTIES ('comment' = '{escape(comment)}');")
        lines.append("")

    # ----------------------------------------------------------------
    # 2. Table-level tags
    # ----------------------------------------------------------------
    table_tags = {}

    # Map workbook fields to tag names
    field_to_tag = [
        ("product name", "product_name"),
        ("product version", "product_version"),
        ("product owner (team)", "product_owner"),
        ("product owner contact", "product_owner_contact"),
        ("grain keys (composite key)", "grain_keys"),
        ("grain keys (composite key — profiler proposed, owner confirmed)", "grain_keys"),
        ("default measure grain (if multiple measures share grain)", "default_measure_grain"),
        ("default measure grain (if all measures share grain)", "default_measure_grain"),
        ("source system", "source_system"),
        ("source table / file", "source_table"),
        ("refresh frequency", "refresh_frequency"),
        ("as-of-date column (snapshot tables only)", "as_of_date_column"),
        ("as-of-date column", "as_of_date_column"),
        ("methodology (default for all measures)", "default_methodology"),
        ("last attested by", "attested_by"),
        ("last attested on", "attested_on"),
        ("contract status", "contract_status"),
        ("expected row count (range)", "expected_row_count"),
        ("data sensitivity / classification", "data_classification"),
    ]

    for field, tag in field_to_tag:
        v = table_meta.get(field, "")
        if v:
            table_tags[tag] = v

    if table_tags:
        lines.append("-- 2. Table-level tags")
        lines.append(f"ALTER TABLE {fqn}")
        lines.append(f"  SET TAGS (")
        items = [f"{indent}'{k}' = '{escape(v)}'" for k, v in table_tags.items()]
        lines.append(",\n".join(items))
        lines.append(f"  );")
        lines.append("")

    # ----------------------------------------------------------------
    # 3. Primary key
    # ----------------------------------------------------------------
    pk_cols = [c["column_name"] for c in columns if c["is_primary_key"].lower() == "true"]
    if pk_cols:
        # Build a constraint name from the table name
        cname = "pk_" + fqn.split(".")[-1].replace("star_", "")[:25]
        lines.append("-- 3. Primary key (informational, NOT ENFORCED)")
        lines.append(f"ALTER TABLE {fqn}")
        lines.append(f"  ADD CONSTRAINT {cname} PRIMARY KEY ({', '.join(pk_cols)}) NOT ENFORCED;")
        lines.append("")

    # ----------------------------------------------------------------
    # 4. Foreign keys
    # ----------------------------------------------------------------
    fks = [c for c in columns if c["is_foreign_key"].lower() == "true"]
    if fks:
        lines.append("-- 4. Foreign keys (informational, NOT ENFORCED)")
        for c in fks:
            ref = c["fk_references"]
            if not ref:
                lines.append(f"-- [WARN] {c['column_name']} marked is_foreign_key=true but fk_references is blank — skipping")
                continue
            # Split "catalog.schema.table.column" → target_table + target_column
            parts = ref.split(".")
            if len(parts) < 4:
                lines.append(f"-- [WARN] {c['column_name']} fk_references not in catalog.schema.table.column form — skipping")
                continue
            target_table = ".".join(parts[:3])
            target_col = parts[3]
            cname = f"fk_{fqn.split('.')[-1].replace('star_', '')[:20]}_{c['column_name'][:20]}"
            lines.append(f"ALTER TABLE {fqn}")
            lines.append(f"  ADD CONSTRAINT {cname}")
            lines.append(f"    FOREIGN KEY ({c['column_name']})")
            lines.append(f"    REFERENCES {target_table}({target_col})")
            lines.append(f"    NOT ENFORCED;")
        lines.append("")

    # ----------------------------------------------------------------
    # 5. Check constraints — allowed values
    # ----------------------------------------------------------------
    allowed_constraints = []
    for c in columns:
        if c["allowed_values"]:
            values = [v.strip() for v in c["allowed_values"].split(",")]
            quoted = ", ".join([f"'{escape(v)}'" for v in values])
            cname = f"chk_{c['column_name'][:30]}_values"
            allowed_constraints.append(
                f"ALTER TABLE {fqn}\n"
                f"  ADD CONSTRAINT {cname}\n"
                f"    CHECK ({c['column_name']} IN ({quoted}));"
            )

    if allowed_constraints:
        lines.append("-- 5a. Check constraints — allowed values")
        lines.append("\n\n".join(allowed_constraints))
        lines.append("")

    # ----------------------------------------------------------------
    # 5b. Check constraints — format / regex
    # ----------------------------------------------------------------
    format_constraints = []
    for c in columns:
        if c["format_regex"]:
            regex = c["format_regex"]
            for prefix in ("Must match ", "Must match", "Matches "):
                if regex.startswith(prefix):
                    regex = regex[len(prefix):].strip()
            # Skip if it doesn't look like a regex (no metacharacters)
            if not any(ch in regex for ch in ("^", "$", "[", "(", "\\", "+", "*", "?", "|", "{")):
                continue
            cname = f"chk_{c['column_name'][:30]}_format"
            format_constraints.append(
                f"ALTER TABLE {fqn}\n"
                f"  ADD CONSTRAINT {cname}\n"
                f"    CHECK ({c['column_name']} RLIKE '{regex}');"
            )

    if format_constraints:
        lines.append("-- 5b. Check constraints — format / regex")
        lines.append("\n\n".join(format_constraints))
        lines.append("")

    # ----------------------------------------------------------------
    # 6. Column comments
    # ----------------------------------------------------------------
    comments = [c for c in columns if c["comment"]]
    if comments:
        lines.append("-- 6. Column comments")
        for c in comments:
            lines.append(f"ALTER TABLE {fqn}")
            lines.append(f"  ALTER COLUMN {c['column_name']} COMMENT '{escape(c['comment'])}';")
        lines.append("")

    # ----------------------------------------------------------------
    # 7. Column-level tags
    # ----------------------------------------------------------------
    lines.append("-- 7. Column-level tags")
    for c in columns:
        col_tags = {}
        if c["role"]:
            col_tags["role"] = c["role"]
        if c["business_name"]:
            col_tags["business_name"] = c["business_name"]
        if c["fk_references"]:
            col_tags["fk_references"] = c["fk_references"]
        if c["fk_cardinality"]:
            col_tags["fk_cardinality"] = c["fk_cardinality"]
        if c["referential_integrity"]:
            col_tags["referential_integrity"] = c["referential_integrity"]
        if c["source_of_truth"]:
            col_tags["source_of_truth"] = c["source_of_truth"]
        if c["measure_grain"]:
            col_tags["measure_grain"] = c["measure_grain"]
        if c["measure_unit"]:
            col_tags["measure_unit"] = c["measure_unit"]
        if c["measure_methodology"]:
            col_tags["measure_methodology"] = c["measure_methodology"]

        # Aggregation rules — parsed from a single field like:
        #   "agg_across_counterparty=sum; agg_across_agreement=sum; ..."
        if c["agg_rules"]:
            for part in c["agg_rules"].split(";"):
                part = part.strip()
                if "=" in part:
                    k, v = part.split("=", 1)
                    col_tags[k.strip()] = v.strip()

        if col_tags:
            lines.append(f"ALTER TABLE {fqn}")
            lines.append(f"  ALTER COLUMN {c['column_name']} SET TAGS (")
            items = [f"{indent}'{k}' = '{escape(v)}'" for k, v in col_tags.items()]
            lines.append(",\n".join(items))
            lines.append(f"  );")
            lines.append("")

    # ----------------------------------------------------------------
    # 8. Verify
    # ----------------------------------------------------------------
    catalog, schema, table = fqn.split(".")
    lines.append("-- " + "=" * 70)
    lines.append("-- Verify — read back the metadata that was applied")
    lines.append("-- " + "=" * 70)
    lines.append("")
    lines.append("-- Table-level tags")
    lines.append("SELECT * FROM system.information_schema.table_tags")
    lines.append(f"WHERE catalog_name = '{catalog}' AND schema_name = '{schema}'")
    lines.append(f"  AND table_name = '{table}';")
    lines.append("")
    lines.append("-- Column-level tags")
    lines.append("SELECT * FROM system.information_schema.column_tags")
    lines.append(f"WHERE catalog_name = '{catalog}' AND schema_name = '{schema}'")
    lines.append(f"  AND table_name = '{table}';")
    lines.append("")
    lines.append("-- Constraints")
    lines.append("SELECT * FROM system.information_schema.table_constraints")
    lines.append(f"WHERE table_catalog = '{catalog}' AND table_schema = '{schema}'")
    lines.append(f"  AND table_name = '{table}';")
    lines.append("")

    return "\n".join(lines)


def escape(s):
    """Escape single quotes for SQL string literals."""
    return s.replace("'", "''")


# ----------------------------------------------------------------------
# Main — generate DDL for both worked examples
# ----------------------------------------------------------------------
def main(workbook_path, output_path):
    wb = load_workbook(workbook_path, data_only=True)

    # Dim
    dim_table = read_table_metadata(wb["Table — Dim Worked Example"])
    dim_cols = read_column_metadata(wb["Column — Dim Worked Example"])
    dim_fqn = "d4001_centralus_tdvip_creditrisk.xvala_xva.star_dim_counterparty"

    # Fact
    fact_table = read_table_metadata(wb["Table — Fact Worked Example"])
    fact_cols = read_column_metadata(wb["Column — Fact Worked Example"])
    fact_fqn = "d4001_centralus_tdvip_creditrisk.xvala_xva.star_fact_issuer_exposure"

    output = []
    output.append("-- " + "#" * 70)
    output.append("-- # DDL Generated from UC Metadata Enrichment Workbook")
    output.append(f"-- # Source: {Path(workbook_path).name}")
    output.append("-- " + "#" * 70)
    output.append("")
    output.append(gen_ddl(dim_fqn, dim_table, dim_cols))
    output.append("")
    output.append(gen_ddl(fact_fqn, fact_table, fact_cols))

    full = "\n".join(output)
    Path(output_path).write_text(full)
    print(f"Generated DDL written to: {output_path}")
    print(f"Lines: {len(full.splitlines())}")
    print(f"Bytes: {len(full)}")


if __name__ == "__main__":
    wb_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/outputs/UC_Metadata_Enrichment_Workbook.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "/home/claude/generated_ddl.sql"
    main(wb_path, out_path)
